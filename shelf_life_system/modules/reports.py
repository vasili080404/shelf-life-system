"""
modules/reports.py
------------------
Модуль Reports — компонент архитектуры прототипа (рис. 2.15).

Содержит формирование отчётов и экспорт в TXT / Excel. Каждый отчёт
представлен dataclass'ом и хелпером для текстовой/табличной выгрузки.
"""
from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from tkinter import filedialog
from typing import Any

from openpyxl import Workbook

from modules.database import Database


# ---------- text formatting ----------
def _row_format(widths: list[int]) -> str:
    return " | ".join(f"{{:>{w}}}" if i % 2 == 1 else f"{{:<{w}}}"
                      for i, w in enumerate(widths))


@dataclass
class ReportTable:
    title: str
    headers: list[str]
    rows: list[list[Any]]
    widths: list[int] | None = None
    footer: list[str] | None = None  # строки после таблицы (например, итоги)

    def to_text(self) -> str:
        widths = self.widths or [
            max(len(str(h)), max((len(str(c)) for r in self.rows for c in r), default=0))
            for h in self.headers
        ]
        fmt = _row_format(widths)
        lines: list[str] = []
        lines.append(self.title)
        lines.append("")
        lines.append(fmt.format(*[str(h) for h in self.headers]))
        lines.append("-" * (sum(widths) + 3 * (len(widths) - 1)))
        for row in self.rows:
            lines.append(fmt.format(*[str(c) for c in row]))
        if self.footer:
            lines.append("-" * (sum(widths) + 3 * (len(widths) - 1)))
            lines.extend(self.footer)
        return "\n".join(lines)

    def to_excel_rows(self) -> list[list[Any]]:
        out: list[list[Any]] = [self.headers, *self.rows]
        if self.footer:
            out.append([])
            for f in self.footer:
                out.append([f])
        return out


# ---------- export helpers ----------
def save_text_report(table: ReportTable, default_name: str) -> str | None:
    today = datetime.now().strftime("%Y-%m-%d")
    path = filedialog.asksaveasfilename(
        defaultextension=".txt",
        initialfile=f"{default_name}_{today}.txt",
        filetypes=[("Text files", "*.txt"), ("All files", "*.*")],
    )
    if not path:
        return None
    Path(path).write_text(table.to_text(), encoding="utf-8")
    return path


def save_excel_report(table: ReportTable, default_name: str) -> str | None:
    today = datetime.now().strftime("%Y-%m-%d")
    path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        initialfile=f"{default_name}_{today}.xlsx",
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
    )
    if not path:
        return None
    wb = Workbook()
    ws = wb.active
    ws.title = default_name[:31] or "Отчёт"
    for r_idx, row in enumerate(table.to_excel_rows(), start=1):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)
    # ширины колонок по максимальной длине
    for c_idx in range(len(table.headers)):
        col_letter = ws.cell(row=1, column=c_idx + 1).column_letter
        max_len = max(
            [len(str(table.headers[c_idx]))]
            + [len(str(r[c_idx])) for r in table.rows if c_idx < len(r)]
        )
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)
    wb.save(path)
    return path


# ---------- high-level report builders ----------
class Reports:
    def __init__(self, db: Database) -> None:
        self.db = db

    def all_products(self) -> ReportTable:
        rows = self.db.get_all_active_batches_with_stock()
        table_rows = [
            [r["ProductName"], r["ExpirationDate"], r["StockQty"]]
            for r in rows
        ]
        return ReportTable(
            title="Список всех товаров в базе данных (FEFO: сначала истекающие)",
            headers=["Название товара", "Срок годности", "Остаток"],
            rows=table_rows,
            widths=[42, 14, 10],
        )

    def sales_history(self, limit: int = 50) -> ReportTable:
        rows = self.db.get_recent_sales(limit)
        table_rows = [
            [r["SaleDate"], r["ProductName"], r["Quantity"]]
            for r in rows
        ]
        return ReportTable(
            title=f"История продаж (последние {limit})",
            headers=["Дата и время", "Товар", "Кол-во"],
            rows=table_rows,
            widths=[20, 42, 8],
        )

    def losses(self, days: int = 7) -> ReportTable:
        rows = self.db.get_losses_in_window(days)
        total_loss = 0.0
        table_rows: list[list[Any]] = []
        for r in rows:
            qty = int(r.get("StockQty") or 0)
            price = float(r.get("PurchasePrice") or 0)
            loss = qty * price
            total_loss += loss
            table_rows.append([
                r["ProductName"], r["ExpirationDate"], qty, f"{price:.2f}", f"{loss:.2f} BYN",
            ])
        return ReportTable(
            title=f"Потери от просрочки за последние {days} дней",
            headers=["Товар", "Срок истёк", "Остаток", "Цена", "Сумма потерь"],
            rows=table_rows,
            widths=[40, 14, 8, 8, 14],
            footer=[f"ИТОГО ПОТЕРЬ ЗА НЕДЕЛЮ: {total_loss:.2f} BYN"],
        )

    def writeoffs(self, limit: int = 100) -> ReportTable:
        rows = self.db.get_writeoffs(limit)
        total_qty = 0
        table_rows: list[list[Any]] = []
        for r in rows:
            qty = int(r.get("Quantity") or 0)
            total_qty += qty
            table_rows.append([
                r.get("WriteOffDate"), r.get("ProductName"), qty,
                r.get("Reason", ""), r.get("ActNumber", ""),
            ])
        return ReportTable(
            title=f"История списаний (последние {limit})",
            headers=["Дата", "Товар", "Кол-во", "Причина", "Акт"],
            rows=table_rows,
            widths=[20, 40, 8, 18, 24],
            footer=[f"ИТОГО СПИСАНО: {total_qty} шт"],
        )

    def my_orders(self) -> ReportTable:
        rows = self.db.list_orders()
        table_rows = [
            [r["OrderID"], r["CreationDate"], r["TotalItems"], r["Status"]]
            for r in rows
        ]
        return ReportTable(
            title="Сохранённые интеллектуальные авто-заказы",
            headers=["№ заказа", "Дата создания", "Позиций", "Статус"],
            rows=table_rows,
            widths=[10, 22, 10, 14],
        )

    def order_details(self, order_id: int) -> ReportTable:
        rows = self.db.get_order_items(order_id)
        table_rows = [
            [r["ProductName"], r["RecommendedQuantity"], r.get("Reason", "")]
            for r in rows
        ]
        return ReportTable(
            title=f"Состав заказа №{order_id}",
            headers=["Товар", "Кол-во", "Причина"],
            rows=table_rows,
            widths=[42, 10, 30],
        )

    def auto_order_proposal(self, recs) -> ReportTable:
        """recs: list[OrderRecommendation] из модуля AutoOrder."""
        table_rows = [
            [r.product_name, r.current_stock, r.avg_daily_sales,
             r.recommended_qty, r.reason]
            for r in recs
        ]
        return ReportTable(
            title="Рекомендуемый интеллектуальный авто-заказ",
            headers=["Товар", "Остаток", "Ср. продажи/день", "Рекомендация", "Причина"],
            rows=table_rows,
            widths=[36, 10, 16, 16, 40],
        )
